"""Spreadsheet parsing and dependency graph construction.

Loads Excel workbooks via openpyxl, extracts cell data and formulas,
builds a networkx dependency graph, and detects structural properties.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import networkx as nx
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.utils.formula_parser import (
    CellReference,
    parse_formula,
    expand_range,
    find_pattern_outliers,
    ERROR_VALUES,
)
from app.models.schemas import SheetMetadata


@dataclass
class CellData:
    """Parsed data for a single cell."""
    sheet: str
    cell_ref: str           # e.g. "B5"
    qualified_ref: str       # e.g. "Sheet1!B5"
    formula: str | None      # None if value-only cell
    value: object            # computed value
    data_type: str           # 'f' = formula, 'n' = number, 's' = string, etc.
    row: int
    col: int
    label: str | None = None  # detected label from adjacent cells


@dataclass
class WorkbookData:
    """Complete parsed workbook data."""
    file_path: str
    sheets: dict[str, list[CellData]] = field(default_factory=dict)
    sheet_metadata: list[SheetMetadata] = field(default_factory=list)
    dependency_graph: nx.DiGraph = field(default_factory=nx.DiGraph)
    all_cells: dict[str, CellData] = field(default_factory=dict)  # qualified_ref -> CellData
    formula_cells: list[CellData] = field(default_factory=list)
    error_cells: list[CellData] = field(default_factory=list)


# ── Sheet type detection patterns ─────────────────────────────────────────────

SHEET_TYPE_PATTERNS = {
    "assumptions": re.compile(r"assump|input|driver|param", re.IGNORECASE),
    "pnl": re.compile(r"p\s*[&n]\s*l|profit\s*(and|&)\s*loss|income\s*state", re.IGNORECASE),
    "cash_flow": re.compile(r"cash\s*flow|cf\b|cash\s*state", re.IGNORECASE),
    "balance_sheet": re.compile(r"balance\s*sheet|bs\b|bal\s*sheet", re.IGNORECASE),
    "cap_table": re.compile(r"cap\s*table|capitali[sz]ation|equity|share|stock", re.IGNORECASE),
    "revenue": re.compile(r"revenue|sales|top\s*line|mrr|arr", re.IGNORECASE),
    "costs": re.compile(r"cost|expense|opex|cogs|overhead", re.IGNORECASE),
    "headcount": re.compile(r"head\s*count|team|hiring|staff|personnel", re.IGNORECASE),
}

# Labels that indicate assumption/metric cells
ASSUMPTION_LABEL_PATTERNS = [
    re.compile(r"growth\s*rate", re.IGNORECASE),
    re.compile(r"churn\s*rate", re.IGNORECASE),
    re.compile(r"cac|customer\s*acquisition\s*cost", re.IGNORECASE),
    re.compile(r"ltv|life\s*time\s*value|clv", re.IGNORECASE),
    re.compile(r"gross\s*margin", re.IGNORECASE),
    re.compile(r"burn\s*rate", re.IGNORECASE),
    re.compile(r"runway", re.IGNORECASE),
    re.compile(r"conversion\s*rate", re.IGNORECASE),
    re.compile(r"arpu|average\s*revenue", re.IGNORECASE),
    re.compile(r"mrr|monthly\s*recurring", re.IGNORECASE),
    re.compile(r"arr|annual\s*recurring", re.IGNORECASE),
    re.compile(r"net\s*margin", re.IGNORECASE),
    re.compile(r"opex|operating\s*expense", re.IGNORECASE),
    re.compile(r"discount\s*rate", re.IGNORECASE),
    re.compile(r"tax\s*rate", re.IGNORECASE),
    re.compile(r"inflation", re.IGNORECASE),
    re.compile(r"retention\s*rate", re.IGNORECASE),
    re.compile(r"payback\s*period", re.IGNORECASE),
]


def parse_workbook(file_path: str | Path) -> WorkbookData:
    """Parse an Excel workbook and build a complete dependency graph.

    Args:
        file_path: Path to the .xlsx file

    Returns:
        WorkbookData with all cells, formulas, dependency graph, and metadata.
    """
    file_path = str(file_path)
    wb_data = WorkbookData(file_path=file_path)

    # Load with formulas preserved (data_only=False)
    wb_formulas = load_workbook(file_path, data_only=False)
    # Also load with computed values
    wb_values = load_workbook(file_path, data_only=True)

    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]

        sheet_cells: list[CellData] = []
        has_formulas = False
        labeled_sections: list[str] = []

        for row in ws_f.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                qualified_ref = f"{sheet_name}!{cell_ref}"

                # Get computed value from data_only workbook
                val_cell = ws_v[cell_ref]
                computed_value = val_cell.value

                formula = None
                data_type = cell.data_type

                if data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    formula = str(cell.value)
                    has_formulas = True
                    data_type = "f"

                # Try to find a label for this cell
                label = _find_cell_label(ws_f, cell.row, cell.column)

                cd = CellData(
                    sheet=sheet_name,
                    cell_ref=cell_ref,
                    qualified_ref=qualified_ref,
                    formula=formula,
                    value=computed_value if computed_value is not None else cell.value,
                    data_type=data_type,
                    row=cell.row,
                    col=cell.column,
                    label=label,
                )

                sheet_cells.append(cd)
                wb_data.all_cells[qualified_ref] = cd

                if formula:
                    wb_data.formula_cells.append(cd)

                # Check for error values
                if isinstance(cd.value, str) and cd.value.upper() in ERROR_VALUES:
                    wb_data.error_cells.append(cd)

        wb_data.sheets[sheet_name] = sheet_cells

        # Build sheet metadata
        sheet_type = _detect_sheet_type(sheet_name, sheet_cells)
        for section_label in _find_labeled_sections(sheet_cells):
            labeled_sections.append(section_label)

        wb_data.sheet_metadata.append(SheetMetadata(
            name=sheet_name,
            row_count=ws_f.max_row or 0,
            col_count=ws_f.max_column or 0,
            has_formulas=has_formulas,
            detected_type=sheet_type,
            labeled_sections=labeled_sections,
        ))

    wb_formulas.close()
    wb_values.close()

    # Build dependency graph
    _build_dependency_graph(wb_data)

    return wb_data


def _find_cell_label(ws, row: int, col: int) -> str | None:
    """Try to find a label for a cell by looking at adjacent cells to the left or above."""
    # Check cell to the left (same row, col-1)
    if col > 1:
        left_cell = ws.cell(row=row, column=col - 1)
        if left_cell.value and isinstance(left_cell.value, str) and left_cell.data_type != "f":
            return str(left_cell.value).strip()

    # Check cell above (row-1, same col)
    if row > 1:
        above_cell = ws.cell(row=row - 1, column=col)
        if above_cell.value and isinstance(above_cell.value, str) and above_cell.data_type != "f":
            return str(above_cell.value).strip()

    return None


def _detect_sheet_type(sheet_name: str, cells: list[CellData]) -> str:
    """Detect the type of a sheet based on its name and content."""
    for type_name, pattern in SHEET_TYPE_PATTERNS.items():
        if pattern.search(sheet_name):
            return type_name

    # Fall back to content-based detection
    text_content = " ".join(
        str(c.value) for c in cells
        if isinstance(c.value, str) and c.data_type != "f"
    )[:2000]

    for type_name, pattern in SHEET_TYPE_PATTERNS.items():
        if pattern.search(text_content):
            return type_name

    return "unknown"


def _find_labeled_sections(cells: list[CellData]) -> list[str]:
    """Find cells that match known financial assumption labels."""
    sections = []
    for cell in cells:
        if cell.label:
            for pattern in ASSUMPTION_LABEL_PATTERNS:
                if pattern.search(cell.label):
                    sections.append(cell.label)
                    break
    return list(set(sections))


def _build_dependency_graph(wb_data: WorkbookData):
    """Build a networkx DiGraph from formula precedents.

    Edges go from precedent → dependent (i.e., A1 → B1 means B1 depends on A1).
    """
    graph = wb_data.dependency_graph

    for cell in wb_data.formula_cells:
        info = parse_formula(cell.formula, cell.sheet)

        # Add the formula cell as a node
        graph.add_node(cell.qualified_ref, **{
            "formula": cell.formula,
            "value": cell.value,
            "sheet": cell.sheet,
        })

        for ref in info.precedent_refs:
            # Resolve sheet name
            ref_sheet = ref.sheet or cell.sheet

            if ref.is_range:
                # Expand range to individual cells
                expanded = expand_range(ref.cell)
                for expanded_cell in expanded:
                    source = f"{ref_sheet}!{expanded_cell}"
                    graph.add_edge(source, cell.qualified_ref)
            else:
                source = f"{ref_sheet}!{ref.cell}"
                graph.add_edge(source, cell.qualified_ref)


def find_circular_references(graph: nx.DiGraph) -> list[list[str]]:
    """Detect circular references in the dependency graph.

    Returns:
        List of cycles, each cycle is a list of qualified cell refs.
    """
    try:
        cycles = list(nx.simple_cycles(graph))
        # Limit to first 50 cycles to avoid explosion
        return cycles[:50]
    except nx.NetworkXError:
        return []


def get_dependents(graph: nx.DiGraph, cell_ref: str) -> list[str]:
    """Get all cells that depend on the given cell (downstream)."""
    if cell_ref not in graph:
        return []
    return list(nx.descendants(graph, cell_ref))


def get_precedents(graph: nx.DiGraph, cell_ref: str) -> list[str]:
    """Get all cells that the given cell depends on (upstream)."""
    if cell_ref not in graph:
        return []
    return list(nx.ancestors(graph, cell_ref))


def find_formula_inconsistencies(wb_data: WorkbookData) -> list[tuple[str, str, str]]:
    """Find cells with formulas that break the pattern of adjacent cells.

    Returns:
        List of (qualified_ref, sheet, reason) tuples.
    """
    inconsistencies = []

    for sheet_name, cells in wb_data.sheets.items():
        # Group cells by row
        rows: dict[int, list[CellData]] = {}
        cols: dict[int, list[CellData]] = {}

        for cell in cells:
            rows.setdefault(cell.row, []).append(cell)
            cols.setdefault(cell.col, []).append(cell)

        # Check each row for pattern breaks
        for row_num, row_cells in rows.items():
            cell_tuples = [(c.cell_ref, c.formula) for c in row_cells if c.formula or c.data_type == "n"]
            outliers = find_pattern_outliers(cell_tuples, sheet_name)
            for cell_ref, reason in outliers:
                inconsistencies.append((f"{sheet_name}!{cell_ref}", sheet_name, reason))

        # Check each column for pattern breaks
        for col_num, col_cells in cols.items():
            cell_tuples = [(c.cell_ref, c.formula) for c in col_cells if c.formula or c.data_type == "n"]
            outliers = find_pattern_outliers(cell_tuples, sheet_name)
            for cell_ref, reason in outliers:
                qual = f"{sheet_name}!{cell_ref}"
                if not any(q == qual for q, _, _ in inconsistencies):
                    inconsistencies.append((qual, sheet_name, reason))

    return inconsistencies
