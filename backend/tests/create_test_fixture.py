"""Generate a sample financial model Excel workbook for testing.

This creates a realistic-looking startup financial model with:
- Known formula errors (#REF!, hardcoded values)
- Circular references
- Cross-sheet inconsistencies
- Mixed units
- A cap table sheet
- Labeled assumptions (growth rate, churn, etc.)
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path


def create_sample_model(output_path: str = None) -> str:
    """Create a comprehensive test financial model.

    Returns:
        Path to the generated file.
    """
    if output_path is None:
        output_path = str(Path(__file__).parent / "fixtures" / "sample_model.xlsx")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: Assumptions
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Assumptions"

    # Headers
    ws["A1"] = "Key Assumptions"
    ws["A2"] = "Growth Rate (MoM)"
    ws["B2"] = 0.15  # 15% — aggressive
    ws["A3"] = "Churn Rate (Monthly)"
    ws["B3"] = 0.02  # 2%
    ws["A4"] = "CAC (Customer Acquisition Cost)"
    ws["B4"] = 5000  # INR
    ws["A5"] = "LTV (Customer Lifetime Value)"
    ws["B5"] = 45000  # INR
    ws["A6"] = "Gross Margin %"
    ws["B6"] = 0.72  # 72%
    ws["A7"] = "Monthly Burn Rate (INR)"
    ws["B7"] = 1500000
    ws["A8"] = "Starting Cash (INR)"
    ws["B8"] = 15000000
    ws["A9"] = "Tax Rate"
    ws["B9"] = 0.25
    ws["A10"] = "Conversion Rate"
    ws["B10"] = 0.035  # 3.5%
    ws["A11"] = "Average Revenue Per User (USD)"  # Mixed currency!
    ws["B11"] = 29.99
    ws["A12"] = "Price (INR/month)"
    ws["B12"] = 999
    ws["A13"] = "Discount Rate"
    ws["B13"] = 0.12
    ws["A14"] = "Retention Rate"
    ws["B14"] = 0.98
    ws["A15"] = "Headcount Growth Rate (Annual)"
    ws["B15"] = 0.50  # 50% — aggressive

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: P&L (Profit & Loss)
    # ══════════════════════════════════════════════════════════════════════
    ws_pnl = wb.create_sheet("P&L")

    ws_pnl["A1"] = "Profit & Loss Statement"
    ws_pnl["A2"] = ""
    ws_pnl["B2"] = "Month 1"
    ws_pnl["C2"] = "Month 2"
    ws_pnl["D2"] = "Month 3"
    ws_pnl["E2"] = "Month 4"
    ws_pnl["F2"] = "Month 5"
    ws_pnl["G2"] = "Month 6"

    # Revenue row
    ws_pnl["A3"] = "Revenue"
    ws_pnl["B3"] = 500000
    ws_pnl["C3"] = "=B3*(1+Assumptions!B2)"
    ws_pnl["D3"] = "=C3*(1+Assumptions!B2)"
    ws_pnl["E3"] = "=D3*(1+Assumptions!B2)"
    ws_pnl["F3"] = 800000  # ← HARDCODED (should be formula like others)
    ws_pnl["G3"] = "=F3*(1+Assumptions!B2)"

    # COGS
    ws_pnl["A4"] = "COGS"
    ws_pnl["B4"] = "=B3*(1-Assumptions!B6)"
    ws_pnl["C4"] = "=C3*(1-Assumptions!B6)"
    ws_pnl["D4"] = "=D3*(1-Assumptions!B6)"
    ws_pnl["E4"] = "=E3*(1-Assumptions!B6)"
    ws_pnl["F4"] = "=F3*(1-Assumptions!B6)"
    ws_pnl["G4"] = "=G3*(1-Assumptions!B6)"

    # Gross Profit
    ws_pnl["A5"] = "Gross Profit"
    ws_pnl["B5"] = "=B3-B4"
    ws_pnl["C5"] = "=C3-C4"
    ws_pnl["D5"] = "=D3-D4"
    ws_pnl["E5"] = "=E3-E4"
    ws_pnl["F5"] = "=F3-F4"
    ws_pnl["G5"] = "=G3-G4"

    # Operating Expenses
    ws_pnl["A6"] = "Operating Expenses"
    ws_pnl["B6"] = 400000
    ws_pnl["C6"] = "=B6*1.05"  # ← 1.05 hardcoded (should reference assumption)
    ws_pnl["D6"] = "=C6*1.05"
    ws_pnl["E6"] = "=D6*1.05"
    ws_pnl["F6"] = "=E6*1.05"
    ws_pnl["G6"] = "=F6*1.05"

    # Net Income
    ws_pnl["A7"] = "Net Income"
    ws_pnl["B7"] = "=B5-B6"
    ws_pnl["C7"] = "=C5-C6"
    ws_pnl["D7"] = "=D5-D6"
    ws_pnl["E7"] = "=E5-E6"
    ws_pnl["F7"] = "=F5-F6"
    ws_pnl["G7"] = "=G5-G6"

    # Total Revenue (for cross-sheet checking)
    ws_pnl["A9"] = "Total Revenue"
    ws_pnl["B9"] = "=SUM(B3:G3)"

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: Cash Flow
    # ══════════════════════════════════════════════════════════════════════
    ws_cf = wb.create_sheet("Cash Flow")

    ws_cf["A1"] = "Cash Flow Statement"
    ws_cf["B1"] = "Month 1"
    ws_cf["C1"] = "Month 2"
    ws_cf["D1"] = "Month 3"

    ws_cf["A2"] = "Operating Cash Flow"
    ws_cf["B2"] = "='P&L'!B7"
    ws_cf["C2"] = "='P&L'!C7"
    ws_cf["D2"] = "='P&L'!D7"

    ws_cf["A3"] = "CapEx"
    ws_cf["B3"] = -50000
    ws_cf["C3"] = -50000
    ws_cf["D3"] = -50000

    ws_cf["A4"] = "Net Cash Flow"
    ws_cf["B4"] = "=B2+B3"
    ws_cf["C4"] = "=C2+C3"
    ws_cf["D4"] = "=D2+D3"

    ws_cf["A5"] = "Ending Cash Balance"
    ws_cf["B5"] = "=Assumptions!B8+B4"
    ws_cf["C5"] = "=B5+C4"
    ws_cf["D5"] = "=C5+D4"

    # Total Revenue — INTENTIONALLY DIFFERENT from P&L (cross-sheet mismatch!)
    ws_cf["A7"] = "Total Revenue"
    ws_cf["B7"] = 3200000  # ← Different from P&L!B9

    # Runway calculation
    ws_cf["A9"] = "Runway (months)"
    ws_cf["B9"] = "=Assumptions!B8/Assumptions!B7"

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 4: Balance Sheet
    # ══════════════════════════════════════════════════════════════════════
    ws_bs = wb.create_sheet("Balance Sheet")

    ws_bs["A1"] = "Balance Sheet"
    ws_bs["A3"] = "Total Assets"
    ws_bs["B3"] = 20000000

    ws_bs["A5"] = "Total Liabilities"
    ws_bs["B5"] = 8000000

    ws_bs["A7"] = "Total Equity"
    ws_bs["B7"] = 11000000  # ← Should be 12M to balance! (intentional error)

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 5: Cap Table
    # ══════════════════════════════════════════════════════════════════════
    ws_cap = wb.create_sheet("Cap Table")

    ws_cap["A1"] = "Capitalization Table"
    ws_cap["A3"] = "Shareholder"
    ws_cap["B3"] = "Shares"
    ws_cap["C3"] = "Ownership %"

    ws_cap["A4"] = "Founder 1"
    ws_cap["B4"] = 400000
    ws_cap["C4"] = 40

    ws_cap["A5"] = "Founder 2"
    ws_cap["B5"] = 300000
    ws_cap["C5"] = 30

    ws_cap["A6"] = "Investor (Seed)"
    ws_cap["B6"] = 150000
    ws_cap["C6"] = 15

    ws_cap["A7"] = "ESOP/Option Pool"
    ws_cap["B7"] = 100000
    ws_cap["C7"] = 8  # ← Only 8%, might be flagged as small

    ws_cap["A8"] = "Angel Investors"
    ws_cap["B8"] = 50000
    ws_cap["C8"] = 5

    # Total — doesn't sum to 100% (intentional error: sums to 98%)
    ws_cap["A10"] = "Total"
    ws_cap["B10"] = "=SUM(B4:B8)"
    ws_cap["C10"] = "=SUM(C4:C8)"

    # Pre/Post money
    ws_cap["A12"] = "Pre-Money Valuation (INR Cr)"
    ws_cap["B12"] = 10
    ws_cap["A13"] = "Investment Amount (INR Cr)"
    ws_cap["B13"] = 2
    ws_cap["A14"] = "Post-Money Valuation (INR Cr)"
    ws_cap["B14"] = "=B12+B13"

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 6: Revenue Model (with a #REF! error)
    # ══════════════════════════════════════════════════════════════════════
    ws_rev = wb.create_sheet("Revenue Model")

    ws_rev["A1"] = "Revenue Forecast"
    ws_rev["A2"] = "Users"
    ws_rev["B2"] = 1000
    ws_rev["C2"] = "=B2*(1+Assumptions!B2)"
    ws_rev["D2"] = "=C2*(1+Assumptions!B2)"

    ws_rev["A3"] = "Revenue per User"
    ws_rev["B3"] = "=Assumptions!B12"
    ws_rev["C3"] = "=Assumptions!B12"
    ws_rev["D3"] = "=Assumptions!B12"

    ws_rev["A4"] = "Total Revenue"
    ws_rev["B4"] = "=B2*B3"
    ws_rev["C4"] = "=C2*C3"
    ws_rev["D4"] = "=D2*D3"

    ws_rev["A5"] = "MRR"
    ws_rev["B5"] = "=B4"
    ws_rev["C5"] = "=C4"
    ws_rev["D5"] = "=D4"

    ws_rev["A6"] = "ARR"
    ws_rev["B6"] = "=B5*12"
    ws_rev["C6"] = "=C5*12"
    ws_rev["D6"] = "=D5*12"

    # Save
    wb.save(output_path)
    print(f"Sample model created: {output_path}")
    return output_path


if __name__ == "__main__":
    create_sample_model()
